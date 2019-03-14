import pylab as plt
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import cmath as cm
from src.tprt import Receiver, Layer, Source, DilatCenter, RotatCenter, Ray, FlatHorizon, GridHorizon, ISOVelocity, Velocity_model
from src.tprt.ray import SnelliusError
from src.tprt.rt_coefficients import rt_coefficients
from src.tprt.bicubic_interpolation import get_left_i

import time # for runtime
import os # for folder's creation
import datetime # for folder's naming

import openpyxl as opxl
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice

from scipy.optimize import minimize

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)

def time_window(gather, rec_time, t_window, t_central):

    if rec_time[0] <= t_central + t_window and rec_time[- 1] >= t_central - t_window:

        i_left = get_left_i(rec_time, t_central - t_window)
        i_right = get_left_i(rec_time, t_central + t_window)


        return gather[i_left : i_right + 1]

    else:

        return np.array([])

start_time = time.time()

# TODO: make class for vel_mod, for now init at least one horizon upper the top receiver
# TODO: check procedure of "is_ray_intersect_boundary"

model_number = 1
curv_scale = 1

# Later it will be needed to insert the current model's name into titles of plots. In order to do this, let's
# define a dictionary:

number_string = {1:'1', 2:'2a', 3:'2b', 4:'2c', 5:'3'}

# We'll calculate the geometrical spreading in several assumptions. For example, we can not consider curvature at
# reflection point, or vice versa, consider the curvature only at this point. So:

transmission_curv = False
reflection_curv = False

if transmission_curv == True and reflection_curv == True:
    reflection_curv == False

# Let's define the interfaces.

# Grid:
if 1 < model_number < 5 or model_number == 6:

    X = np.linspace(- 1200, 1200, 1201)
    Y = np.linspace(- 1200, 1200, 1201)

else:

    X = np.linspace(- 1000, 1000, 1001)
    Y = np.linspace(- 1000, 1000, 1001)

YY, XX = np.meshgrid(Y, X)

# Interfaces:

Convex_Gauss_500_Center = np.array(list(map(lambda x, y: 700 - 200 * np.exp(- x * x / 1000000 * curv_scale -
                                                                            y * y / 1000000 * curv_scale), XX, YY)))
Convex_Gauss_1000_Center = np.array(list(map(lambda x, y: 1200 - 200 * np.exp(- x * x / 1000000 * curv_scale -
                                                                              y * y / 1000000 * curv_scale), XX, YY)))

Convex_Gauss_500_Incident = np.array(list(map(lambda x, y: 700 - 200 * np.exp(- (x + 816.045 * 900 / 1200) *
                                                                              (x + 816.045 * 900 / 1200) / 1000000 * curv_scale -
                                                                              y * y / 1000000 * curv_scale), XX, YY)))
Convex_Gauss_500_Upcoming = np.array(list(map(lambda x, y: 700 - 200 * np.exp(- (x - 816.045 * 900 / 1200) *
                                                                              (x - 816.045 * 900 / 1200) / 1000000 * curv_scale -
                                                                              y * y / 1000000 * curv_scale), XX, YY)))
Concave_Gauss_500_Center = np.array(list(map(lambda x, y: 700 + 200 * np.exp(- x * x / 1000000 * curv_scale -
                                                                             y * y / 1000000 * curv_scale), XX, YY)))
Fourth_Model1 = np.array(list(map(lambda x, y: 3000 - np.sqrt(2500**2 - x**2), XX, YY)))
Fourth_Model2 = np.array(list(map(lambda x, y: 3500 - np.sqrt(2500**2 - x**2), XX, YY)))

# Let's construct raycodes for reflected waves:
raycode_model_1_3 = [[1, 0, 0],
                     [-1, 0, 0]]

raycode_model_2 = [[1, 0, 0],
                   [1, 1, 0],
                   [- 1, 1, 0],
                   [- 1, 0, 0]]

# Let's set a particular model and a particular raycode as "current". The current model will be specified by a number
# (starting with 1) and the current raycode will be set by a rule: model 1 or 3 -> raycode_model_1_3, any kind of
# model 2 -> raycode_model_2.
# model numbers:

# 1 - vel_mod_1
# 2 - vel_mod_2a
# 3 - vel_mod_2b
# 4 - vel_mod_2c
# 5 - vel_mod_3

# Models:

# For IPGG computer:
models_dir_name = "C:/" \
                  "Users/" \
                  "ShilovNN/" \
                  "Documents/" \
                  "Лучевой метод/" \
                  "AVO/" \
                  "Результаты вычислений/" \
                  "Модель №{}/" \
                  "Кривизна {}/"\
                  "Горизонты".format(number_string[model_number], 0.0004 * curv_scale)


# For your laptop:
# models_dir_name = "C:/" \
#                   "Users/" \
#                   "USER/" \
#                   "Documents/" \
#                   "Лучевой метод/" \
#                   "AVO, коэффициенты отражения-преломления/" \
#                   "Результаты вычислений/" \
#                   "Модель №{}/" \
#                   "Кривизна {}/"\
#                   "Горизонты".format(number_string[model_number], 0.0004 * curv_scale)

if model_number == 1:

    first_hor = np.load("{}/Convex_Gauss_500_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Convex_Gauss_500_Center, 0, first_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600)]),
                                 np.array([1800, 2100]), np.array([1, 2]), horizons)

elif model_number == 2:

    first_hor = np.load("{}/Convex_Gauss_500_Incident.npy".format(models_dir_name))

    second_hor = np.load("{}/Convex_Gauss_1000_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Convex_Gauss_500_Incident, 0, first_hor),
                GridHorizon(X, Y, Convex_Gauss_1000_Center, 0, second_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600), ISOVelocity(3600, 2100)]),
                                 np.array([1800, 2100, 2400]), np.array([1, 2, 3]), horizons)

elif model_number == 3:

    first_hor = np.load("{}/Convex_Gauss_500_Center.npy".format(models_dir_name))

    second_hor = np.load("{}/Convex_Gauss_1000_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Convex_Gauss_500_Center, 0, first_hor),
                GridHorizon(X, Y, Convex_Gauss_1000_Center, 0, second_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600), ISOVelocity(3600, 2100)]),
                                 np.array([1800, 2100, 2400]), np.array([1, 2, 3]), horizons)

elif model_number == 4:

    first_hor = np.load("{}/Convex_Gauss_500_Upcoming.npy".format(models_dir_name))

    second_hor = np.load("{}/Convex_Gauss_1000_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Convex_Gauss_500_Upcoming, 0, first_hor),
                GridHorizon(X, Y, Convex_Gauss_1000_Center, 0, second_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600), ISOVelocity(3600, 2100)]),
                                 np.array([1800, 2100, 2400]), np.array([1, 2, 3]), horizons)

elif model_number == 5:

    first_hor = np.load("{}/Concave_Gauss_500_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Concave_Gauss_500_Center, 0, first_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600)]),
                                 np.array([1800, 2100]), np.array([1, 2]), horizons)
elif model_number == 6:

    horizons = [GridHorizon(X, Y, Fourth_Model1, bool_parab = 0),
                GridHorizon(X, Y, Fourth_Model2, bool_parab = 0)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600), ISOVelocity(3600, 2100)]),
                                 np.array([1800, 2100, 2400]), np.array([1, 2, 3]), horizons)

else: # in any oser case set the default model:

    first_hor = np.load("{}/Convex_Gauss_500_Center.npy".format(models_dir_name))

    horizons = [GridHorizon(X, Y, Convex_Gauss_500_Center, 0, first_hor)]
    current_mod = Velocity_model(np.array([ISOVelocity(2000, 1100), ISOVelocity(2800, 1600)]),
                                 np.array([1800, 2100]), np.array([1, 2]), horizons)

# And current raycode:
if 1 < model_number < 5 or model_number == 6:

    current_raycode = raycode_model_2

    refl_i = 1 # number of the reflection interface
    num_segments = 4 # number of segments

else:

    current_raycode = raycode_model_1_3 # by default

    refl_i = 0 # number of the reflection interface
    num_segments = 2 # number of segments

# Let's construct the observation system:
sou_line = np.arange(- 1300, 25, 25) # source line starting from - 1300 and ending at 0 with step 100
rec_line = np.linspace(0, 1300, sou_line.shape[0]) # source line starting from 0 and ending at 1300 with step 100

# Let's initiate all arrays:

# Let's set sources and receivers along profile:
sources = np.empty(sou_line.shape[0], dtype = Source)
receivers = np.empty(rec_line.shape[0], dtype = Receiver)

# So, let's set Sources and Receivers.

frequency_dom = 39 # dominant frequency of the source

for i in range(sou_line.shape[0]):

    sources[i] = DilatCenter(frequency_dom, current_mod, np.array([sou_line[i], 0, 0]))
    receivers[i] = Receiver([rec_line[-1 - i], 0, 0], orientation = np.array([[1, 0, 0],
                                                                              [0, - 1, 0],
                                                                              [0, 0, - 1]]))

# And initiate rays themselves:
rays = np.empty(sou_line.shape[0], dtype = Ray)

# Geometrical spreading along all rays:
geom_spread_curv = np.zeros(rays.shape)
geom_spread_plane = np.zeros(rays.shape)

geom_spread_curv_inv = np.zeros(rays.shape)
geom_spread_plane_inv = np.zeros(rays.shape)

geom_spread_homogen = np.zeros(rays.shape)

# Arrays for reflection / transmission coefficents and for cosines of incidence:

coefficients = np.zeros((rec_line.shape[0], num_segments - 1), dtype = complex)
cosines = np.zeros((rec_line.shape[0], num_segments - 1), dtype = complex)

# Arrays for transformed amplidudes:

transformed_ampl_curv = np.zeros(rec_line.shape[0], dtype = complex)
transformed_ampl_plane = np.zeros(rec_line.shape[0], dtype = complex)

transformed_ampl_homogen = np.zeros(rec_line.shape[0], dtype = complex)

# Travel time of waves from sources to the surface receivers
travel_time = np.zeros(rays.shape)

# Record time at the station in seconds:
max_time = 2

# Time array for seismic gathers. Grid spacing is 1 millisecond:
record_time = np.arange(0, max_time, 0.001)

# A two-dimensional arrays for gathers:
gathers_x = np.zeros((rec_line.shape[0], record_time.shape[0]))
gathers_y = np.zeros((rec_line.shape[0], record_time.shape[0]))
gathers_z = np.zeros((rec_line.shape[0], record_time.shape[0]))

# File management:

# Create a description file. We shall create a folder for this file and others. So, the name of the directory is:

# For IPGG computer:
dir_name = "C:/Users/ShilovNN/Documents/Лучевой метод/AVO/Результаты вычислений/" \
           "Модель №{}/Кривизна {}/Вычисления {} {}-{}".format(number_string[model_number],
                                                               0.0004 * curv_scale,
                                                               datetime.datetime.now().date(),
                                                               datetime.datetime.now().hour,
                                                               datetime.datetime.now().minute)

# For your laptop:
# dir_name = "C:/Users/USER/Documents/Лучевой метод/AVO, коэффициенты отражения-преломления/Результаты вычислений/" \
#            "Модель №{}/Кривизна {}/Вычисления {} {}-{}".format(number_string[model_number],
#                                                                0.0004 * curv_scale,
#                                                                datetime.datetime.now().date(),
#                                                                datetime.datetime.now().hour,
#                                                                datetime.datetime.now().minute)

createFolder(dir_name)

description_file = open("{}/Описание.txt".format(dir_name), "w+")

description_file.write("Модель №{} \n \n".format(number_string[model_number]))
description_file.write("Сетка задния границ: \n \n")
description_file.write("Узлы на оси X расположены от {} до {} м с шагом {} м. \n".format(X[0],
                                                                             X[- 1],
                                                                             X[1] - X[0]))
description_file.write("Узлы на оси Y расположены от {} до {} м с шагом {} м. \n \n".format(Y[0],
                                                                             Y[- 1],
                                                                             Y[1] - Y[0]))
description_file.write("Схема наблюдений: \n \n")
description_file.write("Источники расположены на оси X от {} до {} м с шагом {} м. \n".format(int(sou_line[0]),
                                                                                              int(sou_line[- 1]),
                                                                                              int(sou_line[1] -
                                                                                                  sou_line[0])))
description_file.write("Приёмники расположены на оси X от {} до {} м с шагом {} м. \n \n".format(int(rec_line[0]),
                                                                                                 int(rec_line[- 1]),
                                                                                                 int(rec_line[1] -
                                                                                                     rec_line[0])))
description_file.write("Функция источника: импульс Рикера с главной частотой {} Гц. \n \n".format(frequency_dom))
description_file.write("Для пар 'источник-приёмник', находящихся в закритической области,"
                       " лучи строиться не будут.\n\n")
description_file.write("Время записи на станции: {} с \n".format(max_time))
description_file.write("Длительность отсчётов: {} с \n \n".format(record_time[1] - record_time[0]))

description_file.write("В графиках и численных данных с пометкой <<без учёта кривизны границ>> не учитывается: \n\n")

if 1 < model_number < 5 and transmission_curv == False:
    description_file.write("- Кривизна преломляющих границ;")
if reflection_curv == False:
    description_file.write("\n- Кривизна отражающей границы;")

description_file.write("\n\n")

description_file.write("============================================================ \n \n")

description_file.write("Время вычислений: \n \n")
description_file.write("Под обработкой луча понимается расчёт амплитуд и геометрического расхождения. \n")
description_file.write("Сетка задния границ: \n \n")
description_file.write("--- Конфигурации границ и схемы наблюдений заданы: %s секунд --- \n \n" % (time.time() -
                                                                                                   start_time))
print("\x1b[1;31m --- The configuration of interfaces and the observation system are defined: %s seconds ---\n" %
      (time.time() - start_time))

# Constructing rays:

for i in range(sources.shape[0]):

    print("------------------------------------")
    rays[i] = Ray(sources[- 1 - i], receivers[- 1 - i], current_mod, current_raycode)
    rays[i].optimize(snells_law = True, dtravel = True)

    description_file.write("--- %s луч создан и оптимизирован: %s секунд --- \n" % (i + 1, time.time() - start_time))
    print("\x1b[1;31m --- %s ray constructed: %s seconds ---" % (i + 1, time.time() - start_time))

    rays[i].amplitude_fun, coefficients[i, :], cosines[i, :] = rays[i].amplitude_fr_dom() # rewrite the amplitude field.

    # Check if we are in post-critical zone:

    if np.linalg.norm(np.imag(rays[i].amplitude_fun))!= 0:

        rays = np.delete(rays, np.arange(i, rays.shape[0], 1), axis = 0)

        geom_spread_curv = np.delete(geom_spread_curv, np.arange(i, geom_spread_curv.shape[0], 1), axis = 0)
        geom_spread_plane = np.delete(geom_spread_plane, np.arange(i, geom_spread_plane.shape[0], 1), axis = 0)

        geom_spread_curv_inv = np.delete(geom_spread_curv_inv, np.arange(i, geom_spread_curv_inv.shape[0], 1), axis = 0)
        geom_spread_plane_inv = np.delete(geom_spread_plane_inv, np.arange(i, geom_spread_plane_inv.shape[0], 1), axis = 0)

        geom_spread_homogen = np.delete(geom_spread_homogen, np.arange(i, geom_spread_homogen.shape[0], 1), axis = 0)

        coefficients = np.delete(coefficients, np.arange(i, coefficients.shape[0], 1), axis = 0)
        cosines = np.delete(cosines, np.arange(i, cosines.shape[0], 1), axis = 0)

        transformed_ampl_curv = np.delete(transformed_ampl_curv, np.arange(i, transformed_ampl_curv.shape[0], 1), axis = 0)
        transformed_ampl_plane = np.delete(transformed_ampl_plane, np.arange(i, transformed_ampl_plane.shape[0], 1), axis = 0)

        transformed_ampl_homogen = np.delete(transformed_ampl_homogen,
                                             np.arange(i, transformed_ampl_homogen.shape[0], 1), axis = 0)

        travel_time = np.delete(travel_time, np.arange(i, travel_time.shape[0], 1), axis = 0)

        gathers_x = np.delete(gathers_x, np.arange(i, gathers_x.shape[0], 1), axis = 0)
        gathers_y = np.delete(gathers_y, np.arange(i, gathers_y.shape[0], 1), axis = 0)
        gathers_z = np.delete(gathers_z, np.arange(i, gathers_z.shape[0], 1), axis = 0)

        description_file.write("--- На %s луче был достигнут критический угол. Процесс остановлен: %s секунд --- \n\n" %
                               (i + 1, time.time() - start_time))
        print("\x1b[1;31m --- On the %s-th ray critical angle has been reached. The process has been terminated:"
              " %s seconds ---" % (i + 1, time.time() - start_time))

        break
    if np.linalg.norm(np.imag(coefficients[i, :])) != 0:

        print(np.imag(coefficients[i, :]))

    travel_time[i] = rays[i].travel_time()

    geom_spread_curv[i] = rays[i].spreading([True, True], 0)
    geom_spread_plane[i] = rays[i].spreading([transmission_curv, reflection_curv], 0)

    geom_spread_curv_inv[i] = rays[i].spreading([True, True], 1)
    geom_spread_plane_inv[i] = rays[i].spreading([transmission_curv, reflection_curv], 1)

    geom_spread_homogen[i] = (np.linalg.norm(rays[i].segments[0].source -
                                            np.array([0, 0, horizons[- 1].get_depth([0, 0])])) +
                              np.linalg.norm(rays[i].segments[- 1].receiver -
                                             np.array([0, 0, horizons[- 1].get_depth([0, 0])]))) ** 2

    for j in range(record_time.shape[0]):

        gathers_x[i, j], gathers_y[i, j], gathers_z[i, j] = np.real(rays[i].amplitude_t_dom(record_time[j]))

    description_file.write("--- %s луч обработан: %s секунд --- \n \n" % (i + 1, time.time() - start_time))
    print("\x1b[1;31m --- %s ray processed: %s seconds ---" % (i + 1, time.time() - start_time))

print("------------------------------------")

# print("main curvatures =",current_mod.horizons[0].get_sec_deriv(rays[-1].segments[-2].receiver[0:2],rays[-1].segments[-2].vector)[0:3])
# print("point of incidence =",rays[-1].segments[-2].receiver)
# print("point of reflection =",rays[-1].segments[-2].source)
# print("main curvatures at the refl. point =",current_mod.horizons[-1].get_sec_deriv(rays[-1].segments[-2].source[0:2],rays[-1].segments[-3].vector)[0:3])

# Files filling:

# It will be convenient to create .xlsx files where we shall write values of the
# amplitude, traveltime and geometrical spreading.

geom_spread = opxl.Workbook()
seismogram_x = opxl.Workbook()
seismogram_y = opxl.Workbook()
seismogram_z = opxl.Workbook()
ray_amplitude = opxl.Workbook()
hodograph = opxl.Workbook()

geom_spread_sheet = geom_spread.active
geom_spread_sheet.title = "Геометрическое расхождение"

seismogram_x_sheet = seismogram_x.active
seismogram_y_sheet = seismogram_y.active
seismogram_z_sheet = seismogram_z.active
seismogram_x_sheet.title = "X-компонента"
seismogram_y_sheet.title = "Y-компонента"
seismogram_z_sheet.title = "Z-компонента"

ray_amplitude_sheet = ray_amplitude.active
ray_amplitude_sheet.title = "Лучевые амплитуды"

hodograph_sheet = hodograph.active
hodograph_sheet.title = "Годограф ОСТ"

# Let's form up heads for these files:
geom_spread_sheet.cell(row = 1, column = 1).value = "Геометрическое расхождение"

geom_spread_sheet.cell(row = 3, column = 1).value = "X, м"
geom_spread_sheet.cell(row = 3, column = 2).value = "С учётом кривизны границы, м^2"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == False:
            geom_spread_sheet.cell(row = 3, column = 3).value = "Без учёта кривизны границ, м^2"
        else:
            geom_spread_sheet.cell(row = 3, column = 3).value = "Без учёта кривизны преломляющих границ, м^2"
    else:
        geom_spread_sheet.cell(row = 3, column = 3).value = "Без учёта кривизны отражающей границы, м^2"
else:
    if reflection_curv == False:
        geom_spread_sheet.cell(row = 3, column = 3).value = "Без учёта кривизны отражающей границы, м^2"
geom_spread_sheet.cell(row = 3, column = 4).value = "Без учёта преломляющих границ и кривизны в точке отражения, м^2"

seismogram_x_sheet.cell(row = 1, column = 1).value = "X-компонента вектора смещений"
seismogram_y_sheet.cell(row = 1, column = 1).value = "Y-компонента вектора смещений"
seismogram_z_sheet.cell(row = 1, column = 1).value = "Z-компонента вектора смещений"

seismogram_x_sheet.cell(row = 3, column = 1).value = "T, с"
seismogram_y_sheet.cell(row = 3, column = 1).value = "T, с"
seismogram_z_sheet.cell(row = 3, column = 1).value = "T, с"

ray_amplitude_sheet.cell(row = 1, column = 1).value = "Лучевые амплитуды"
ray_amplitude_sheet.cell(row = 3, column = 1).value = "Координата приёмника, м"
ray_amplitude_sheet.cell(row = 3, column = 2).value =  "Амплитуда, у.е."

for i in range(receivers.shape[0]):

    seismogram_x_sheet.cell(row = 3, column = i + 2).value = "Тр. №{}".format(i + 1)
    seismogram_y_sheet.cell(row = 3, column = i + 2).value = "Тр. №{}".format(i + 1)
    seismogram_z_sheet.cell(row = 3, column = i + 2).value = "Тр. №{}".format(i + 1)

hodograph_sheet.cell(row = 1, column = 1).value = "Годограф первых вступлений"
hodograph_sheet.cell(row = 3, column = 1).value = "X, м"
hodograph_sheet.cell(row = 3, column = 2).value = "T, с"

# Now, let's save seismograms:

max_amplitude = max(np.max(abs(gathers_z)), np.max(abs(gathers_x)))
accuracy = int(round(abs(np.log10(max_amplitude))) * 5)

for i in range(record_time.shape[0]):

    seismogram_x_sheet.cell(row = 3 + 1 + i, column = 1).value = round(record_time[i], 3)
    seismogram_y_sheet.cell(row = 3 + 1 + i, column = 1).value = round(record_time[i], 3)
    seismogram_z_sheet.cell(row = 3 + 1 + i, column = 1).value = round(record_time[i], 3)

    for j in range(rays.shape[0]):

        seismogram_x_sheet.cell(row = 3 + 1 + i, column = 2 + j).value = round(gathers_x[j, i], accuracy)
        seismogram_y_sheet.cell(row = 3 + 1 + i, column = 2 + j).value = round(gathers_y[j, i], accuracy)
        seismogram_z_sheet.cell(row = 3 + 1 + i, column = 2 + j).value = round(gathers_z[j, i], accuracy)

# And other data:
for i in range(rays.shape[0]):

    ray_amplitude_sheet.cell(row = 3 + 1 + i, column = 1).value = rec_line[i]
    ray_amplitude_sheet.cell(row = 3 + 1 + i, column = 2).value = float(np.linalg.norm(rays[i].amplitude_fun))

    geom_spread_sheet.cell(row = 3 + 1 + i, column = 1).value = rec_line[i]
    geom_spread_sheet.cell(row = 3 + 1 + i, column = 2).value = geom_spread_curv[i]
    geom_spread_sheet.cell(row = 3 + 1 + i, column = 3).value = geom_spread_plane[i]
    geom_spread_sheet.cell(row = 3 + 1 + i, column = 4).value = geom_spread_homogen[i]

    hodograph_sheet.cell(row = 3 + 1 + i, column = 1).value = rec_line[i]
    hodograph_sheet.cell(row = 3 + 1 + i, column = 2).value = travel_time[i]

geom_spread.save("{}/Геометрическое расхождение.xlsx".format(dir_name))
seismogram_x.save("{}/Сейсмограммы. X-компонента.xlsx".format(dir_name))
seismogram_y.save("{}/Сейсмограммы. Y-компонента.xlsx".format(dir_name))
seismogram_z.save("{}/Сейсмограммы. Z-компонента.xlsx".format(dir_name))
ray_amplitude.save("{}/Лучевые амплитуды.xlsx".format(dir_name))
hodograph.save("{}/Годограф.xlsx".format(dir_name))

geom_spread.close()
seismogram_x.close()
seismogram_y.close()
seismogram_z.close()
ray_amplitude.close()
hodograph.close()

# The seismograms in the inversion and adjacent procedures will be noisy. The distribution of the noise should be normal
# with zero mean value. Its dispersion will be proportional to the average absolute value in the Ricker wavelet. So,
# let's form up these noisy gathers:

average_signal = np.zeros(rays.shape[0])
for i in range(rays.shape[0]):

    average_signal[i] = np.average(time_window(np.sqrt(gathers_x[i] ** 2
                                                       + gathers_y[i] ** 2
                                                       + gathers_z[i] ** 2),
                                               record_time,
                                               3 * 1.5 * 1 / frequency_dom,
                                               travel_time[i]))



noise_dispersion = 0.1 * np.max(average_signal)

gathers_x_inv = gathers_x + np.random.randn(gathers_x.shape[0], gathers_x.shape[1]) * noise_dispersion
gathers_y_inv = gathers_y + np.random.randn(gathers_y.shape[0], gathers_y.shape[1]) * noise_dispersion
gathers_z_inv = gathers_z + np.random.randn(gathers_z.shape[0], gathers_z.shape[1]) * noise_dispersion

# There will be no nedd in saving these arrays. So, we move on!

description_file.write("============================================================ \n\n")
description_file.write("Файлы с численными данными сохранены."
                       " Переход к построению графиков: {} секунд\n\n".format(time.time() - start_time))
print("\nData files have been saved. Plotting process has been started: {} seconds\n".format(time.time() - start_time))

# Plotting:
# Plot rays and the medium:

fig = plt.figure()
ax = Axes3D(fig)
ax.invert_zaxis()

ax.view_init(0, - 90)

for l in current_mod.layers[:-1]:
    l.bottom.plot(ax=ax)

for i in range(rays.shape[0]):

    sources[- 1 - i].plot(ax=ax, color='r', marker='p', s=50)
    receivers[- 1 - i].plot(ax=ax, color='k', marker='^', s=50)

    rays[i].plot(ax=ax)

if model_number == 1 or model_number == 5:

    ax.set_xlim3d(- 1000, 1000)
    ax.set_ylim3d(- 1000, 1000)

if 1< model_number < 5:

    ax.set_xlim3d(- 1200, 1200)
    ax.set_ylim3d(- 1200, 1200)

# Create cubic bounding box to simulate equal aspect ratio
max_range = np.array([horizons[-1].X.max()-horizons[-1].X.min(), horizons[-1].Y.max()-horizons[-1].Y.min(), horizons[-1].Z.max()-horizons[-1].Z.min()]).max()
Xb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][0].flatten() + 0.5*(horizons[-1].X.max()+horizons[-1].X.min())
Yb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][1].flatten() + 0.5*(horizons[-1].Y.max()+horizons[-1].Y.min())
Zb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][2].flatten() + 0.5*(horizons[-1].Z.max()+horizons[-1].Z.min())
# Comment or uncomment following both lines to test the fake bounding box:
for xb, yb, zb in zip(Xb, Yb, Zb):
    ax.plot([xb], [yb], [zb], 'w')

ax.set_xlabel("Расстояние по оси x, м")
# ax.set_ylabel("Расстояние по оси y, м",)
ax.set_zlabel("Глубина, м")

plt.savefig("{}/Лучи.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("Лучевая схема сохранена: {} секунд".format(time.time() - start_time))

plt.close(fig)

fig2 = plt.figure()

plt.title("Геометрическое расхождение. Модель №{}".format(number_string[model_number]))

title = "Без учёта кривизны границ"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == True:
            title = "Без учёта кривизны преломляющих границ"
    else:
        title = "Без учёта кривизны отражающей границы"
else:
    if reflection_curv == False:
        title = "Без учёта кривизны отражающей границы"

plt.plot(rec_line[0 : geom_spread_curv.shape[0]], geom_spread_curv, 'r-',
         label = "С учётом кривизны границ")
plt.plot(rec_line[0 : geom_spread_plane.shape[0]], geom_spread_plane, 'r--',
         label = title)

plt.legend()
plt.grid()

plt.xlabel("Координаты вдоль профиля, м")
plt.ylabel("Геометрическое расхождение, м^2")

plt.savefig("{}/Геометрическое расхождение.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("График геометрического расхождения сохранён: {} секунд".format(time.time() - start_time))

plt.close(fig2)

figN = plt.figure()

plt.title("Геометрическое расхождение на инверсию. Модель №{}".format(number_string[model_number]))

title = "Без учёта кривизны границ"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == True:
            title = "Без учёта кривизны преломляющих границ"
    else:
        title = "Без учёта кривизны отражающей границы"
else:
    if reflection_curv == False:
        title = "Без учёта кривизны отражающей границы"

plt.plot(rec_line[0 : geom_spread_curv_inv.shape[0]], geom_spread_curv_inv, 'r-',
         label = "С учётом кривизны границ")
plt.plot(rec_line[0 : geom_spread_plane_inv.shape[0]], geom_spread_plane_inv, 'r--',
         label = title)

plt.legend()
plt.grid()

plt.xlabel("Координаты вдоль профиля, м")
plt.ylabel("Геометрическое расхождение на инверсию, м^2")

plt.savefig("{}/Геометрическое расхождение на инверсию.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("График геометрического расхождения на инверсию сохранён: {} секунд".format(time.time() - start_time))

plt.close(figN)

fig4 = plt.figure()

plt.title("Годограф. Модель №{}".format(number_string[model_number]))

plt.plot(rec_line[0 : travel_time.shape[0]], travel_time, 'r-', label = "Годограф отражённой PP-волны")

plt.legend()
plt.grid()

plt.xlabel("Координаты вдоль профиля, м")
plt.ylabel("Время первых вступлений, с")

plt.savefig("{}/Годограф ОСТ.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("Годограф сохранён: {} секунд".format(time.time() - start_time))

plt.close(fig4)

fig5 = plt.figure()

plt.title("Сейсмограмма (X-компонента). Модель №{}".format(number_string[model_number]))
plt.gca().invert_yaxis()
plt.yticks(np.arange(0, rays.shape[0], 1), ["{}".format(int(j)) for j in rec_line[0 : rays.shape[0]]],
           fontsize = - 5 / 16 * rays.shape[0] + 11 + 21 * 5 / 16)

for i in range(rays.shape[0]):

    plt.fill_between(record_time, gathers_x_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                     np.ones(record_time.shape) * i,
                     linewidth = 0.3, color = 'r', alpha = 0.5)

    plt.fill_between(time_window(record_time,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     time_window(gathers_x_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     np.ones(time_window(record_time,
                                         record_time,
                                         3 * 1.5 * 1 / frequency_dom,
                                         travel_time[i]).shape) * i,
                     linewidth = 0.3, color = 'g', alpha = 0.5)

plt.ylabel("Координаты вдоль профиля, м")
plt.xlabel("Время, с")

plt.savefig("{}/Сейсмограмма. X-компонента.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("Сейсмограмма (X-компонента) сохранена: {} секунд".format(time.time() - start_time))

plt.close(fig5)

fig6 = plt.figure()

plt.title("Сейсмограмма (Y-компонента). Модель №{}".format(number_string[model_number]))
plt.gca().invert_yaxis()

plt.yticks(np.arange(0, rays.shape[0], 1), ["{}".format(int(j)) for j in rec_line[0 : rays.shape[0]]],
           fontsize = - 5 / 16 * rays.shape[0] + 11 + 21 * 5 / 16)

for i in range(rays.shape[0]):

    plt.fill_between(record_time, gathers_y_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                     np.ones(record_time.shape) * i,
                     linewidth = 0.3, color = 'g', alpha = 0.5)

    plt.fill_between(time_window(record_time,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     time_window(gathers_y_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     np.ones(time_window(record_time,
                                         record_time,
                                         3 * 1.5 * 1 / frequency_dom,
                                         travel_time[i]).shape) * i,
                     linewidth = 0.3, color = 'r', alpha = 0.5)


plt.ylabel("Координаты вдоль профиля, м")
plt.xlabel("Время, с")

plt.savefig("{}/Сейсмограмма. Y-компонента.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("Сейсмограмма (Y-компонента) сохранена: {} секунд".format(time.time() - start_time))

plt.close(fig6)

fig7 = plt.figure()

plt.title("Сейсмограмма (Z-компонента). Модель №{}".format(number_string[model_number]))
plt.gca().invert_yaxis()

plt.yticks(np.arange(0, rays.shape[0], 1), ["{}".format(int(j)) for j in rec_line[0 : rays.shape[0]]],
           fontsize = - 5 / 16 * rays.shape[0] + 11 + 21 * 5 / 16)

for i in range(rays.shape[0]):

    plt.fill_between(record_time, gathers_z_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                     np.ones(record_time.shape) * i,
                     linewidth = 0.3, color = 'b', alpha = 0.5)

    plt.fill_between(time_window(record_time,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     time_window(gathers_z_inv[i, :] / np.max(abs(gathers_z)) / 1.5 + i,
                                 record_time,
                                 3 * 1.5 * 1 / frequency_dom,
                                 travel_time[i]),
                     np.ones(time_window(record_time,
                                         record_time,
                                         3 * 1.5 * 1 / frequency_dom,
                                         travel_time[i]).shape) * i,
                     linewidth = 0.3, color = 'g', alpha = 0.5)


plt.ylabel("Координаты вдоль профиля, м")
plt.xlabel("Время, с")

plt.savefig("{}/Сейсмограмма. Z-компонента.png".format(dir_name), dpi = 400, bbox_inches = 'tight')
print("Сейсмограмма (Z-компонента) сохранена: {} секунд\n".format(time.time() - start_time))

plt.close(fig7)

description_file.write("Графики построены и сохранены. Запускается процедура инверсии: %s секунд\n\n" % (time.time() - start_time))
print("\x1b[1;31mPlots have been saved. Inversion procedure has been started: {} seconds\n".format((time.time() - start_time)))

# Let's go ahead to the inversion.
inversion = opxl.Workbook()

# We shall work with noisy data. Since the noise is random, results of the inversion will also be random. In order to
# obtain more stable results we shall perform several iterations of the inversion algorithms, at each step creating new
# noisy gathers. Let's set up total number of iterations:
number_of_iterations = 1


# Let's introduce a simple function which will return PP-reflection coefficient based on elastic properties of the
# medium and angle of incidence. It will just execute exact formula from "Quantitative Seismology" by Aki, Richards.

def simple_pp_refl_coeff(vp1, vs1, rho1, vp2, vs2, rho2, cos_inc):

    sin_inc = np.sqrt(1 - cos_inc**2)

    sin_refl_p = sin_inc * vp1 / vp1
    sin_refl_s = sin_inc * vs1 / vp1

    cos_refl_p = cm.sqrt(1 - sin_refl_p ** 2)
    cos_refl_s = cm.sqrt(1 - sin_refl_s ** 2)

    sin_tr_p = sin_inc * vp2 / vp1
    sin_tr_s = sin_inc * vs2 / vp1

    cos_tr_p = cm.sqrt(1 - sin_tr_p ** 2)
    cos_tr_s = cm.sqrt(1 - sin_tr_s ** 2)

    #     в данном источнике используются следующие коэффициенты:

    p = sin_inc / vp1 #это параметр луча

    a = rho2 / 1000 * (1 - 2 * (vs2**2) * (p**2)) - rho1 / 1000 * (1 - 2 * (vs1**2) * (p**2))
    b = rho2 / 1000 * (1 - 2 * (vs2**2) * (p**2)) + 2 * rho1 / 1000 * (vs1**2) * (p**2)
    c = rho1 / 1000 * (1 - 2 * (vs1**2) * (p**2)) + 2 * rho2 / 1000 * (vs2**2) * (p**2)
    d = 2 * (rho2 / 1000 * (vs2**2) - rho1 / 1000 * (vs1**2))

    E = b * cos_refl_p / vp1 + c * cos_tr_p / vp2
    F = b * cos_refl_s / vs1 + c * cos_tr_s / vs2
    G = a - d * (cos_refl_p / vp1) * (cos_tr_s / vs2)
    H = a - d * (cos_tr_p / vp2) * (cos_refl_s / vs1)

    D = E * F + G * H * (p**2)

    Rp = ((b * cos_refl_p / vp1 - c * cos_tr_p / vp2) * F -
          (a + d * (cos_refl_p / vp1) * (cos_tr_s / vs2)) * H * (p**2)) / D

    return Rp

# cff = np.zeros(50)
# coss = np.linspace(1,0.4,50)
#
# for i in range(50):
#
#     cff[i] = simple_pp_refl_coeff(2800, 1400, 2500, 4000, 2000, 2500, coss[i])
#
# plt.plot(np.degrees(np.arccos(coss)), cff)
# plt.show()


# Let's define the target functional:
def AVO_residual(layer_2_params, layer_1_params, real_coeff, real_cosines):

    # vp1 = layer_1_params[0]
    # vs1 = layer_1_params[1]
    # rho1 = layer_1_params[2]
    #
    # vp2 = layer_2_params[0]
    # vs2 = layer_2_params[1]
    # rho2 = layer_2_params[2]

    synt_coeff = np.zeros(real_coeff.shape, dtype = complex)

    for i in range(synt_coeff.shape[0]):

        synt_coeff[i] = simple_pp_refl_coeff(layer_1_params[0], layer_1_params[1], layer_1_params[2],
                                             layer_2_params[0], layer_2_params[1], layer_2_params[2],
                                             real_cosines[i])

    return np.linalg.norm(synt_coeff - real_coeff)


# And an auxiliary function:
def RMS(gather, rec_time, t_window, t_central):

    if rec_time[0] <= t_central + t_window and rec_time[- 1] >= t_central - t_window:

        i_left = get_left_i(rec_time, t_central - t_window)
        i_right = get_left_i(rec_time, t_central + t_window)

        return np.linalg.norm(gather[i_left : i_right + 1]) / np.sqrt(gather[i_left : i_right + 1].shape[0])

    else:

        return 0


# Initial guess for the minimizer:
vp_init = current_mod.layers[- 1].get_velocity(0)["vp"] * 1.15
vs_init = current_mod.layers[- 1].get_velocity(0)["vs"] * 0.85
rho_init = current_mod.layers[- 1].get_density() * 1.1


# We shall need to write and save some data during these procedures. Consequently, it would be convenient to create all
# corresponding arrays at one moment.

# Let's create an array for cosines of angles of incidence for calculations for homogeneous overburden:
cosines_homogen = np.zeros(rays.shape)

# Let's introduce arrays where we shall collect all transformed data:
transformed_ampl_curv_array = np.zeros((number_of_iterations, rays.shape[0]))
transformed_ampl_plane_array = np.zeros((number_of_iterations, rays.shape[0]))
transformed_ampl_homogen_array = np.zeros((number_of_iterations, rays.shape[0]))

# Let's get started!
for n in range(number_of_iterations):

    # Let's add some random noise to our seismograms:

    gathers_x_inv = gathers_x + np.random.randn(gathers_x.shape[0], gathers_x.shape[1]) * noise_dispersion
    gathers_y_inv = gathers_y + np.random.randn(gathers_y.shape[0], gathers_y.shape[1]) * noise_dispersion
    gathers_z_inv = gathers_z + np.random.randn(gathers_z.shape[0], gathers_z.shape[1]) * noise_dispersion

    for i in range(rays.shape[0]):

        # Let's find RMS of the amplitude. We'll sum squared amplitudes in a window with width of 3 * 1.5 * T where
        # T = 1 / frequency_dom.

        transformed_ampl_curv[i] = RMS(np.sqrt(gathers_x_inv[i] ** 2 + gathers_z_inv[i] ** 2),
                                       record_time,
                                       3 * 1.5 * 1 / frequency_dom,
                                       travel_time[i]) # we assume that there is only noise at the Y component.

        # transformed_ampl_curv[i] = np.linalg.norm(rays[i].amplitude_fun)

        # transformed_ampl_curv[i] = RMS(np.sqrt(gathers_x_inv[i] ** 2 + gathers_y_inv[i] ** 2 + gathers_z_inv[i] ** 2),
        #                                record_time,
        #                                3 * 1.5 * 1 / frequency_dom,
        #                                travel_time[i])

        transformed_ampl_plane[i] = transformed_ampl_curv[i]
        transformed_ampl_homogen[i] = transformed_ampl_curv[i]

        cosines_homogen[i] = horizons[- 1].get_depth([0, 0]) /\
                             np.linalg.norm(rays[i].segments[0].source -
                                            np.array([0, 0, horizons[- 1].get_depth([0, 0])]))

        # Cancel all transmission coefficients from the amplitudes:
        for j in range(coefficients.shape[1]):

            if j != refl_i:

                transformed_ampl_curv[i] = transformed_ampl_curv[i] / coefficients[i, j]
                transformed_ampl_plane[i] = transformed_ampl_plane[i] / coefficients[i, j]

        # And add geometrical spreading factor:
        transformed_ampl_curv[i] = transformed_ampl_curv[i] * np.sqrt(geom_spread_curv_inv[i])
        transformed_ampl_plane[i] = transformed_ampl_plane[i] * np.sqrt(geom_spread_plane_inv[i])
        transformed_ampl_homogen[i] = transformed_ampl_homogen[i] * np.sqrt(geom_spread_homogen[i])

    # We are interested in variation of the amplitude along the profile. So, at zero offset all amplitudes should be
    # equal to the zero-angle reflection amplitude (we assume that zero-angle reflection coefficient is known):

    transformed_ampl_curv = transformed_ampl_curv /\
                            transformed_ampl_curv[0] *\
                            rt_coefficients(current_mod.layers[refl_i].get_velocity(0)["vp"],
                                            current_mod.layers[refl_i].get_velocity(0)["vs"],
                                            current_mod.layers[refl_i].get_density(),
                                            current_mod.layers[refl_i + 1].get_velocity(0)["vp"],
                                            current_mod.layers[refl_i + 1].get_velocity(0)["vs"],
                                            current_mod.layers[refl_i + 1].get_density(),
                                            1,
                                            np.array([0, 0, 1]),
                                            current_mod.layers[refl_i].get_velocity(0)['vp'],
                                            - 1)[0]

    transformed_ampl_plane = transformed_ampl_plane  /\
                             transformed_ampl_plane[0] *\
                             rt_coefficients(current_mod.layers[refl_i].get_velocity(0)["vp"],
                                             current_mod.layers[refl_i].get_velocity(0)["vs"],
                                             current_mod.layers[refl_i].get_density(),
                                             current_mod.layers[refl_i + 1].get_velocity(0)["vp"],
                                             current_mod.layers[refl_i + 1].get_velocity(0)["vs"],
                                             current_mod.layers[refl_i + 1].get_density(),
                                             1,
                                             np.array([0, 0, 1]),
                                             current_mod.layers[refl_i].get_velocity(0)['vp'],
                                             - 1)[0]

    transformed_ampl_homogen = transformed_ampl_homogen /\
                               transformed_ampl_homogen[0] *\
                               rt_coefficients(current_mod.layers[refl_i].get_velocity(0)["vp"],
                                               current_mod.layers[refl_i].get_velocity(0)["vs"],
                                               current_mod.layers[refl_i].get_density(),
                                               current_mod.layers[refl_i + 1].get_velocity(0)["vp"],
                                               current_mod.layers[refl_i + 1].get_velocity(0)["vs"],
                                               current_mod.layers[refl_i + 1].get_density(),
                                               1,
                                               np.array([0, 0, 1]),
                                               current_mod.layers[refl_i].get_velocity(0)['vp'],
                                               - 1)[0]


    # Fill arrays for .xlsx files. Note that we are working in the pre-critical area, so np.real cancels zero
    # imaginary part.
    transformed_ampl_curv_array[n] = np.real(transformed_ampl_curv)
    transformed_ampl_plane_array[n] = np.real(transformed_ampl_plane)
    transformed_ampl_homogen_array[n] = np.real(transformed_ampl_homogen)

    # Let's minimize our functionals:
    minim_result_curv = minimize(AVO_residual,
                                 np.array([3000, 1500, 2100]),
                                 args = (np.array([current_mod.layers[refl_i].get_velocity(0)["vp"],
                                                   current_mod.layers[refl_i].get_velocity(0)["vs"],
                                                   current_mod.layers[refl_i].get_density()]),
                                         transformed_ampl_curv,
                                         cosines[:, refl_i])).x

    minim_result_plane = minimize(AVO_residual,
                                  np.array([3000, 1500, 2100]),
                                  args = (np.array([current_mod.layers[refl_i].get_velocity(0)["vp"],
                                                    current_mod.layers[refl_i].get_velocity(0)["vs"],
                                                    current_mod.layers[refl_i].get_density()]),
                                          transformed_ampl_plane,
                                          cosines[:, refl_i])).x

    minim_result_homogen = minimize(AVO_residual,
                                    np.array([3000, 1500, 2100]),
                                    args = (np.array([current_mod.layers[refl_i].get_velocity(0)["vp"],
                                                      current_mod.layers[refl_i].get_velocity(0)["vs"],
                                                      current_mod.layers[refl_i].get_density()]),
                                            transformed_ampl_homogen,
                                            cosines_homogen)).x


    # Now, let's write the results in .xlsx file:
    if n == 0:
        inversion_sheet = inversion.active
        inversion_sheet.title = "Инверсия. Итерация №1"
    else:
        inversion_sheet = inversion.create_sheet("Инверсия. Итерация №{}".format(n + 1))


    inversion_sheet.cell(row = 1, column = 1).value = "Данные AVO-инверсии"
    inversion_sheet.cell(row = 3, column = 1).value = "Восстанавливаются характеристики слоя №{}".format(refl_i + 2)
    inversion_sheet.cell(row = 4, column = 1).value = "В сейсмограммы независимо внесены погрешности 10% от cреднего" \
                                                      " значения амплитуд в импульсе"
    inversion_sheet.cell(row = 5, column = 1).value = "Значения параметров"
    inversion_sheet.cell(row = 6, column = 2).value = "Модель"
    inversion_sheet.cell(row = 6, column = 3).value = "Начальное приближение"
    inversion_sheet.cell(row = 6, column = 4).value = "Инверсия c корректно учтённым геометрическим расхождением"
    if 1 < model_number < 5:
        if transmission_curv == False:
            if reflection_curv == False:
                inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                                  "кривизны границ"
            else:
                inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                                  "кривизны преломляющих границ"
        else:
            inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                              "кривизны отражающей границы"
    else:
        if reflection_curv == False:
            inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                              "кривизны отражающей границы"
    inversion_sheet.cell(row = 6, column = 6).value = "Инверсия c геометрическим расхождением без учёта преломляющих " \
                                                      "границ и кривизны в точке отражения"

    inversion_sheet.cell(row = 7, column = 1).value = "Vp, м/с"
    inversion_sheet.cell(row = 8, column = 1).value = "Vs, м/с"
    inversion_sheet.cell(row = 9, column = 1).value = "Dens, кг/м^3"

    inversion_sheet.cell(row = 7, column = 2).value = float(current_mod.layers[refl_i + 1].get_velocity(0)['vp'])
    inversion_sheet.cell(row = 7, column = 3).value = float(vp_init)
    inversion_sheet.cell(row = 7, column = 4).value = float(minim_result_curv[0])
    inversion_sheet.cell(row = 7, column = 5).value = float(minim_result_plane[0])
    inversion_sheet.cell(row = 7, column = 6).value = float(minim_result_homogen[0])

    inversion_sheet.cell(row = 8, column = 2).value = float(current_mod.layers[refl_i + 1].get_velocity(0)['vs'])
    inversion_sheet.cell(row = 8, column = 3).value = float(vs_init)
    inversion_sheet.cell(row = 8, column = 4).value = float(minim_result_curv[1])
    inversion_sheet.cell(row = 8, column = 5).value = float(minim_result_plane[1])
    inversion_sheet.cell(row = 8, column = 6).value = float(minim_result_homogen[1])

    inversion_sheet.cell(row = 9, column = 2).value = float(current_mod.layers[refl_i + 1].get_density())
    inversion_sheet.cell(row = 9, column = 3).value = float(rho_init)
    inversion_sheet.cell(row = 9, column = 4).value = float(minim_result_curv[2])
    inversion_sheet.cell(row = 9, column = 5).value = float(minim_result_plane[2])
    inversion_sheet.cell(row = 9, column = 6).value = float(minim_result_homogen[2])

    inversion_sheet.cell(row = 11, column = 1).value = "Относительная погрешность в процентах"
    inversion_sheet.cell(row = 12, column = 2).value = "Инверсия c корректно учтённым геометрическим расхождением"
    if 1 < model_number < 5:
        if transmission_curv == False:
            if reflection_curv == False:
                inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                                   "кривизны границ"
            else:
                inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                                   "кривизны преломляющих границ"
        else:
            inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                               "кривизны отражающей границы"
    else:
        if reflection_curv == False:
            inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                               "кривизны отражающей границы"
    inversion_sheet.cell(row = 12, column = 4).value = "Инверсия c геометрическим расхождением без учёта преломляющих " \
                                                       "границ и кривизны в точке отражения"

    inversion_sheet.cell(row = 13, column = 1).value = "delta Vp, %"
    inversion_sheet.cell(row = 14, column = 1).value = "delta Vs, %"
    inversion_sheet.cell(row = 15, column = 1).value = "delta Dens, %"

    inversion_sheet.cell(row = 13, column = 2).value = float(abs(minim_result_curv[0] -
                                                                 current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)
    inversion_sheet.cell(row = 13, column = 3).value = float(abs(minim_result_plane[0] -
                                                                 current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)
    inversion_sheet.cell(row = 13, column = 4).value = float(abs(minim_result_homogen[0] -
                                                                 current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)

    inversion_sheet.cell(row = 14, column = 2).value = float(abs(minim_result_curv[1] -
                                                                 current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)
    inversion_sheet.cell(row = 14, column = 3).value = float(abs(minim_result_plane[1] -
                                                                 current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)
    inversion_sheet.cell(row = 14, column = 4).value = float(abs(minim_result_homogen[1] -
                                                                current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                            current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)

    inversion_sheet.cell(row = 15, column = 2).value = float(abs(minim_result_curv[2] -
                                                                 current_mod.layers[refl_i + 1].get_density()) /
                                                             current_mod.layers[refl_i + 1].get_density() * 100)
    inversion_sheet.cell(row = 15, column = 3).value = float(abs(minim_result_plane[2] -
                                                                 current_mod.layers[refl_i + 1].get_density()) /
                                                             current_mod.layers[refl_i + 1].get_density() * 100)
    inversion_sheet.cell(row = 15, column = 4).value = float(abs(minim_result_homogen[2] -
                                                                 current_mod.layers[refl_i + 1].get_density()) /
                                                             current_mod.layers[refl_i + 1].get_density() * 100)


# Now all inversions are finished, but we'd like to automatically take average from all of them.
inversion_sheet = inversion.create_sheet("Инверсия. Среднее")

inversion_sheet.cell(row = 1, column = 1).value = "Усреднение данных AVO-инверсии"
inversion_sheet.cell(row = 3, column = 1).value = "Восстанавливаются характеристики слоя №{}".format(refl_i + 2)
inversion_sheet.cell(row = 4, column = 1).value = "В сейсмограммы независимо внесены погрешности 10% от cреднего" \
                                                  " значения амплитуд в импульсе"
inversion_sheet.cell(row = 5, column = 1).value = "Значения параметров"
inversion_sheet.cell(row = 6, column = 2).value = "Модель"
inversion_sheet.cell(row = 6, column = 3).value = "Начальное приближение"
inversion_sheet.cell(row = 6, column = 4).value = "Инверсия c корректно учтённым геометрическим расхождением"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == False:
            inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                              "кривизны границ"
        else:
            inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                              "кривизны преломляющих границ"
    else:
        inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                          "кривизны отражающей границы"
else:
    if reflection_curv == False:
        inversion_sheet.cell(row = 6, column = 5).value = "Инверсия c геометрическим расхождением без учёта " \
                                                          "кривизны отражающей границы"
inversion_sheet.cell(row = 6, column = 6).value = "Инверсия c геометрическим расхождением без учёта преломляющих " \
                                                  "границ и кривизны в точке отражения"

inversion_sheet.cell(row = 7, column = 1).value = "Vp, м/с"
inversion_sheet.cell(row = 8, column = 1).value = "Vs, м/с"
inversion_sheet.cell(row = 9, column = 1).value = "Dens, кг/м^3"

inversion_sheet.cell(row = 7, column = 2).value = float(current_mod.layers[refl_i + 1].get_velocity(0)['vp'])
inversion_sheet.cell(row = 7, column = 3).value = float(vp_init)
inversion_sheet.cell(row = 8, column = 2).value = float(current_mod.layers[refl_i + 1].get_velocity(0)['vs'])
inversion_sheet.cell(row = 8, column = 3).value = float(vs_init)
inversion_sheet.cell(row = 9, column = 2).value = float(current_mod.layers[refl_i + 1].get_density())
inversion_sheet.cell(row = 9, column = 3).value = float(rho_init)

averageVp = np.zeros(3)
averageVs = np.zeros(3)
averageDens = np.zeros(3)

for i in range(number_of_iterations):

    sheet = inversion["Инверсия. Итерация №{}".format(i + 1)]

    averageVp = averageVp + np.array([sheet["D7"].value, sheet["E7"].value, sheet["F7"].value])
    averageVs = averageVs + np.array([sheet["D8"].value, sheet["E8"].value, sheet["F8"].value])
    averageDens = averageDens + np.array([sheet["D9"].value, sheet["E9"].value, sheet["F9"].value])

averageVp = averageVp / number_of_iterations
averageVs = averageVs / number_of_iterations
averageDens = averageDens / number_of_iterations

inversion_sheet.cell(row = 7, column = 4).value = averageVp[0]
inversion_sheet.cell(row = 8, column = 4).value = averageVs[0]
inversion_sheet.cell(row = 9, column = 4).value = averageDens[0]


inversion_sheet.cell(row = 7, column = 5).value = averageVp[1]
inversion_sheet.cell(row = 8, column = 5).value = averageVs[1]
inversion_sheet.cell(row = 9, column = 5).value = averageDens[1]


inversion_sheet.cell(row = 7, column = 6).value = averageVp[2]
inversion_sheet.cell(row = 8, column = 6).value = averageVs[2]
inversion_sheet.cell(row = 9, column = 6).value = averageDens[2]

inversion_sheet.cell(row = 11, column = 1).value = "Относительная погрешность в процентах"
inversion_sheet.cell(row = 12, column = 2).value = "Инверсия c корректно учтённым геометрическим расхождением"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == False:
            inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                               "кривизны границ"
        else:
            inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                               "кривизны преломляющих границ"
    else:
        inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                           "кривизны отражающей границы"
else:
    if reflection_curv == False:
        inversion_sheet.cell(row = 12, column = 3).value = "Инверсия c геометрическим расхождением без учёта " \
                                                           "кривизны отражающей границы"
inversion_sheet.cell(row = 12, column = 4).value = "Инверсия c геометрическим расхождением без учёта преломляющих " \
                                                   "границ и кривизны в точке отражения"

inversion_sheet.cell(row = 13, column = 1).value = "delta Vp, %"
inversion_sheet.cell(row = 14, column = 1).value = "delta Vs, %"
inversion_sheet.cell(row = 15, column = 1).value = "delta Dens, %"

inversion_sheet.cell(row = 13, column = 2).value = float(abs(averageVp[0] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)
inversion_sheet.cell(row = 13, column = 3).value = float(abs(averageVp[1] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)
inversion_sheet.cell(row = 13, column = 4).value = float(abs(averageVp[2] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vp']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vp'] * 100)

inversion_sheet.cell(row = 14, column = 2).value = float(abs(averageVs[0] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)
inversion_sheet.cell(row = 14, column = 3).value = float(abs(averageVs[1] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)
inversion_sheet.cell(row = 14, column = 4).value = float(abs(averageVs[2] -
                                                             current_mod.layers[refl_i + 1].get_velocity(0)['vs']) /
                                                         current_mod.layers[refl_i + 1].get_velocity(0)['vs'] * 100)

inversion_sheet.cell(row = 15, column = 2).value = float(abs(averageDens[0] -
                                                             current_mod.layers[refl_i + 1].get_density()) /
                                                         current_mod.layers[refl_i + 1].get_density() * 100)
inversion_sheet.cell(row = 15, column = 3).value = float(abs(averageDens[1] -
                                                             current_mod.layers[refl_i + 1].get_density()) /
                                                         current_mod.layers[refl_i + 1].get_density() * 100)
inversion_sheet.cell(row = 15, column = 4).value = float(abs(averageDens[2] -
                                                             current_mod.layers[refl_i + 1].get_density()) /
                                                         current_mod.layers[refl_i + 1].get_density() * 100)

description_file.write("Инверсия завершена. Переход к сохранению "
                       "трансформированных амплитуд: {} секунд\n\n".format((time.time() - start_time)))
print("\x1b[1;31mThe inversion has been finished. Now saving"
      " transformed amplitudes: {} seconds\n".format((time.time() - start_time)))

# And finally, let's close the inversion file:
inversion.save("{}/Данные инверсии.xlsx".format(dir_name))
inversion.close()


# Time has come to write into .xlsx files all transformed amplitudes.
transformed_ampl = opxl.Workbook()

# Create a sheet for correct transformation (curvatures are taken into account):
curved_sheet = transformed_ampl.active
curved_sheet.title = "С учётом кривизны"

# Create a sheet for incorrect transformation (curvatures are not taken into account):
plane_sheet = transformed_ampl.create_sheet("Без учёта кривизны")
# Create a sheet for incorrect transformation (spherical divergence):
homogen_sheet = transformed_ampl.create_sheet("Сферическое расхождение")
# Create a sheet for actual reflection coefficients:
real_sheet = transformed_ampl.create_sheet("Коэффициенты отражения")


# Fill the .xlsx file:
curved_sheet["A1"].value = "Значения трансформированных среднеквадратичных амплитуд"
curved_sheet["A3"].value = "В геометрическом расхождении учитывалась кривизна всех границ"
curved_sheet["A5"].value = "№ итерации \ X, м"

plane_sheet["A1"].value = "Значения трансформированных среднеквадратичных амплитуд"
if 1 < model_number < 5:
    if transmission_curv == False:
        if reflection_curv == False:
            plane_sheet["A3"].value = "В геометрическом расхождении не учитывалась кривизна всех границ"
        else:
            plane_sheet["A3"].value = "В геометрическом расхождении не учитывалась кривизна преломляющих границ"
    else:
        plane_sheet["A3"].value = "В геометрическом расхождении не учитывалась кривизна отражающей границы"
else:
    if reflection_curv == False:
        plane_sheet["A3"].value = "В геометрическом расхождении не учитывалась кривизна отражающей границы"
plane_sheet["A5"].value = "№ итерации \ X, м"

homogen_sheet["A1"].value = "Значения трансформированных среднеквадратичных амплитуд"
homogen_sheet["A3"].value = "В геометрическом расхождении не учитывались преломляющие слои и кривизна отражающей" \
                            " границы"
homogen_sheet["A5"].value = "№ итерации \ X, м"

real_sheet["A1"].value = "Значения коэффициентов отражения"
real_sheet["A3"].value = "X, м"
real_sheet["A4"].value = "Коэффициент"

for i in range(number_of_iterations):
    curved_sheet.cell(row = 6 + i, column = 1).value = i + 1
    plane_sheet.cell(row = 6 + i, column = 1).value = i + 1
    homogen_sheet.cell(row = 6 + i, column = 1).value = i + 1
    for j in range(rays.shape[0]):

        curved_sheet.cell(row = 6 + i, column = j + 2).value = transformed_ampl_curv_array[i, j]
        plane_sheet.cell(row = 6 + i, column = j + 2).value = transformed_ampl_plane_array[i, j]
        homogen_sheet.cell(row = 6 + i, column = j + 2).value = transformed_ampl_homogen_array[i, j]

for i in range(rays.shape[0]):

    curved_sheet.cell(row = 5, column = i + 2).value = rec_line[i]
    plane_sheet.cell(row = 5, column = i + 2).value = rec_line[i]
    homogen_sheet.cell(row = 5, column = i + 2).value = rec_line[i]

    real_sheet.cell(row = 3, column = i + 2).value = rec_line[i]
    real_sheet.cell(row = 4, column = i + 2).value = np.real(coefficients[i, refl_i])

# Let's create a chart representing all curves of transformated amplitudes:
chart_sheet = transformed_ampl.create_sheet("График")

chart = ScatterChart()
chart.title = "Коэффициенты отражения"
chart.style = 9
chart.x_axis.title = "X, м"
chart.y_axis.title = "Коэффициент"

x_values = Reference(real_sheet, min_col = 2, max_col = rays.shape[0] + 1, min_row = 3)

for i in range(number_of_iterations):

    curv_values = Reference(curved_sheet, min_col = 2, max_col = rays.shape[0] + 1, min_row = 6 + i)
    plane_values = Reference(plane_sheet, min_col=2, max_col=rays.shape[0] + 1, min_row=6 + i)
    homogen_values = Reference(homogen_sheet, min_col=2, max_col=rays.shape[0] + 1, min_row=6 + i)

    curv_series = Series(curv_values, x_values, title = "С учётом кривизны. Итерация №{}".format(i + 1))
    plane_series = Series(plane_values, x_values, title="Без учёта кривизны. Итерация №{}".format(i + 1))
    homogen_series = Series(homogen_values, x_values, title="Сферическое расхождение. Итерация №{}".format(i + 1))

    curv_Prop = LineProperties(solidFill = ColorChoice(prstClr='lightCoral'))
    plane_Prop = LineProperties(solidFill = ColorChoice(prstClr='ltGreen'))
    homogen_Prop = LineProperties(solidFill = ColorChoice(prstClr='skyBlue'))

    curv_series.graphicalProperties.line = curv_Prop
    plane_series.graphicalProperties.line = plane_Prop
    homogen_series.graphicalProperties.line = homogen_Prop

    curv_series.graphicalProperties.line.width = 18000  # it's 0.5 mm in EMUs
    plane_series.graphicalProperties.line.width = 18000  # it's 0.5 mm in EMUs
    homogen_series.graphicalProperties.line.width = 18000  # it's 0.5 mm in EMUs

    chart.series.append(curv_series)
    chart.series.append(plane_series)
    chart.series.append(homogen_series)


# Add there a curve of the actual coefficients:
real_values = Reference(real_sheet, min_col = 2, max_col = rays.shape[0] + 1, min_row = 4)
real_series = Series(real_values, x_values, title = "Истинные коэффициенты")

real_Prop = LineProperties(solidFill = ColorChoice(prstClr='black'))
real_series.graphicalProperties.line = real_Prop
real_series.graphicalProperties.line.width = 36000  # it's 1 mm in EMUs

chart.series.append(real_series)

chart_sheet.add_chart(chart, "A1")


# Finally, save it:
transformed_ampl.save("{}/Трансформированные амплитуды.xlsx".format(dir_name))
transformed_ampl.close()

description_file.write("Трансформированные амплитуды сохранены. Переход к отрисовке разрезов функционала невязки: "
                       "{} секунд\n\n".format((time.time() - start_time)))
print("\x1b[1;31mTransformed amplitudes have been saved. Now plotting cross-sections of the residual functional:"
      " {} seconds\n".format((time.time() - start_time)))


# Let's form up 2D sections of the residual functional. For this purpose we shall create special forlder:
createFolder("{}/Срезы функционала невязки".format(dir_name))

# In addition, we would like to create special folders for inversions with correct and incorrect geometrical spreading:
createFolder("{}/Срезы функционала невязки/Геометрическое расхождение с учётом кривизн".format(dir_name))
createFolder("{}/Срезы функционала невязки/Геометрическое расхождение без учёта кривизн".format(dir_name))
createFolder("{}/Срезы функционала невязки/Сферическое расхождение".format(dir_name))

# We want to construct 3 + 3 + 3 cross-sections. Then, we need to set up three values of Vp, three values of Vs and
# three values of rho defining corresponding planes. Central values will be actual ones. Two others will add or subtract
# 20% of it.

vp_planes = np.array([current_mod.layers[refl_i + 1].get_velocity(0)["vp"] * 0.8,
                      current_mod.layers[refl_i + 1].get_velocity(0)["vp"],
                      current_mod.layers[refl_i + 1].get_velocity(0)["vp"] * 1.2])

vs_planes = np.array([current_mod.layers[refl_i + 1].get_velocity(0)["vs"] * 0.8,
                      current_mod.layers[refl_i + 1].get_velocity(0)["vs"],
                      current_mod.layers[refl_i + 1].get_velocity(0)["vs"] * 1.2])

rho_planes = np.array([current_mod.layers[refl_i + 1].get_density() * 0.8,
                      current_mod.layers[refl_i + 1].get_density(),
                      current_mod.layers[refl_i + 1].get_density() * 1.2])

param_planes = np.array([vp_planes, vs_planes, rho_planes])

# In order to perform all calculations and plotting procedures automatically, let's set up dictionaries:
spread_type = {0 : "Геометрическое расхождение с учётом кривизн",
               1 : "Геометрическое расхождение без учёта кривизн",
               2 : "Сферическое расхождение"}
const_param = {0 : "Vp", 1 : "Vs", 2 : "Rho"}
param_units = {0: "м/с", 1 : "м/с", 2 : "кг/м^3"}

# Now let's go forward to cross-section construction and plotting.
cross_section = np.zeros((100, 100))

for n in range(3):  # n defines constant parameter

    # We need to come up with two remaining parameters which can vary
    param_indices = np.delete(np.array([0, 1, 2]), n)

    # Now we have to understand in which order we should pass our parameters to the functional.
    pass_ind = np.argsort(np.array([param_indices[0], param_indices[1], n]))

    # Let's define the grid:

    first_param = np.linspace(param_planes[param_indices[0], 0], param_planes[param_indices[0], 2], 100)
    sec_param = np.linspace(param_planes[param_indices[1], 0], param_planes[param_indices[1], 2], 100)

    for k in range(3): # k defines current value of the constant parameter

        for m in range(3):  # m defines type of the geometrical spreading in data

            curr_spread = np.array([transformed_ampl_curv, transformed_ampl_plane, transformed_ampl_homogen])[m]

            # Finally, we can construct our cross-section:

            for i in range(first_param.shape[0]):
                for j in range(sec_param.shape[0]):
                    # Let's construct an array of arguments in correct order for the functional:

                    arg_param = np.array([first_param[i], sec_param[j], param_planes[n, k]])[pass_ind]

                    if arg_param[0] / arg_param[1] >= np.sqrt(2):

                        cross_section[i, j] = AVO_residual(arg_param,
                                                           np.array([current_mod.layers[refl_i].get_velocity(0)["vp"],
                                                                     current_mod.layers[refl_i].get_velocity(0)["vs"],
                                                                     current_mod.layers[refl_i].get_density()]),
                                                           curr_spread,
                                                           cosines[:, refl_i])

                    else:

                        cross_section[i, j] = np.nan


            # And plot it:

            fig8 = plt.figure()

            plt.title("Cрез функционала невязки в плоскости {} = {} {}\n"
                      "{}".format(const_param[n], param_planes[n, k], param_units[n], spread_type[m]))

            plt.contourf(first_param, sec_param, cross_section.T,
                         levels=10,
                         cmap="magma")
            plt.colorbar()

            plt.contour(first_param, sec_param, cross_section.T,
                        levels=10,
                        colors="gray")

            plt.xlabel("{},{}".format(const_param[param_indices[0]], param_units[param_indices[0]]))
            plt.ylabel("{},{}".format(const_param[param_indices[1]], param_units[param_indices[1]]))

            plt.savefig("{}/"
                        "Срезы функционала невязки/"
                        "{}/"
                        "Cрез функционала невязки в плоскости {} = {}.png".format(dir_name,
                                                                                  spread_type[m],
                                                                                  const_param[n],
                                                                                  param_planes[n, k]),
                        dpi=400, bbox_inches = 'tight')

            plt.close(fig8)


description_file.write("Срезы функционала невязки нарисованы и сохранены: "
                       "{} секунд\n\n".format((time.time() - start_time)))
print("\x1b[1;31mCross-sections of the residual functional are plotted and saved:"
      " {} seconds\n".format((time.time() - start_time)))

description_file.close()
# the_horizons = opxl.Workbook()
# the_horizons_sheet = the_horizons.active
# the_horizons_sheet.title = "Horizon №1"
#
# the_horizons_sheet.cell(row = 1, column = 1).value = "Таблица значений глубины для горизонта №1"
# the_horizons_sheet.cell(row = 3, column = 1).value = "X \ Y"
#
# for i in range(current_mod.horizons[0].X.shape[0]):
#
#     the_horizons_sheet.cell(row = 4 + i, column = 1).value = current_mod.horizons[0].X[i]
#
# for j in range(current_mod.horizons[0].Y.shape[0]):
#
#     the_horizons_sheet.cell(row = 3, column = 2 + j).value = current_mod.horizons[0].Y[j]
#
# for i in range(current_mod.horizons[0].X.shape[0]):
#     for j in range(current_mod.horizons[0].Y.shape[0]):
#
#         the_horizons_sheet.cell(row = 4 + i, column = 2 + j).value = current_mod.horizons[0].Z[i, j]
#
#
# for k in np.arange(1, len(current_mod.horizons), 1):
#
#     the_horizons_sheet = the_horizons.create_sheet("Horizon №{}".format(k + 1))
#
#     the_horizons_sheet.cell(row = 1, column = 1).value = "Таблица значений глубины для горизонта №{}".format(k + 1)
#
#     the_horizons_sheet.cell(row = 3, column = 1).value = "X \ Y"
#
#     for i in range(current_mod.horizons[k].X.shape[0]):
#
#         the_horizons_sheet.cell(row = 4 + i, column = 1).value = current_mod.horizons[k].X[i]
#
#     for j in range(current_mod.horizons[k].Y.shape[0]):
#
#         the_horizons_sheet.cell(row = 3, column = 2 + j).value = current_mod.horizons[k].Y[j]
#
#     for i in range(current_mod.horizons[k].X.shape[0]):
#         for j in range(current_mod.horizons[k].Y.shape[0]):
#
#             the_horizons_sheet.cell(row = 4 + i, column = 2 + j).value = current_mod.horizons[k].Z[i, j]
#
# the_horizons.save("{}/Horizons.xlsx".format(dir_name))
# the_horizons.close()
#
# the_rays = opxl.Workbook()
# the_rays_sheet = the_rays.active
# the_rays_sheet.title = "Rays"
#
# the_rays_sheet.cell(row = 1, column = 1).value = "Rays"
#
# the_rays_sheet.cell(row = 4, column = 1).value = "X{}".format(1)
# the_rays_sheet.cell(row = 5, column = 1).value = "Y{}".format(1)
# the_rays_sheet.cell(row = 6, column = 1).value = "Z{}".format(1)
#
# for k in range(rays.shape[0]):
#
#     the_rays_sheet.cell(row = 3, column = k + 2).value = "Луч №{}".format(k)
#
#     the_rays_sheet.cell(row = 4, column = k + 2).value = float(rays[k].segments[0].source[0])
#     the_rays_sheet.cell(row = 5, column = k + 2).value = float(rays[k].segments[0].source[1])
#     the_rays_sheet.cell(row = 6, column = k + 2).value = float(rays[k].segments[0].source[2])
#
# for i in np.arange(1, len(rays[0].segments) + 1, 1):
#
#     the_rays_sheet.cell(row = 3 * i + 4, column = 1).value = "X{}".format(i + 1)
#     the_rays_sheet.cell(row = 3 * i + 5, column = 1).value = "Y{}".format(i + 1)
#     the_rays_sheet.cell(row = 3 * i + 6, column = 1).value = "Z{}".format(i + 1)
#
#     for k in range(rays.shape[0]):
#
#         the_rays_sheet.cell(row = 3 * i + 4, column = k + 2).value = float(rays[k].segments[i - 1].receiver[0])
#         the_rays_sheet.cell(row = 3 * i + 5, column = k + 2).value = float(rays[k].segments[i - 1].receiver[1])
#         the_rays_sheet.cell(row = 3 * i + 6, column = k + 2).value = float(rays[k].segments[i - 1].receiver[2])
#
#
# the_rays.save("{}/Rays.xlsx".format(dir_name))
# the_rays.close()

# if model_number == 6:
#
#     theor_spread = np.zeros(rays.shape[0])
#
#     for i in range(rays.shape[0]):
#
#         l1 = np.linalg.norm(rays[i].segments[0].get_distance())
#         l2 = np.linalg.norm(rays[i].segments[1].get_distance())
#         l3 = np.linalg.norm(rays[i].segments[2].get_distance())
#         l4 = np.linalg.norm(rays[i].segments[3].get_distance())
#
#         v1 = rays[i].segments[0].layer.get_velocity(0)[rays[i].segments[0].vtype]
#         v2 = rays[i].segments[1].layer.get_velocity(0)[rays[i].segments[1].vtype]
#         v3 = rays[i].segments[2].layer.get_velocity(0)[rays[i].segments[2].vtype]
#         v4 = rays[i].segments[3].layer.get_velocity(0)[rays[i].segments[3].vtype]
#
#         l = np.array([l1, l2, l3, l4])
#         v = np.array([v1, v2, v3, v4])
#
#         l_primes = np.zeros(4)
#         l_double_primes = np.zeros(4)
#
#         for j in range(3):
#
#             cos_inc = abs(np.dot(rays[i].segments[j].vector, rays[i].segments[j].end_horizon.
#                                  get_normal(rays[i].segments[j].receiver[0:2])))
#             cos_out = abs(np.dot(rays[i].segments[j + 1].vector, rays[i].segments[j].end_horizon.
#                                  get_normal(rays[i].segments[j].receiver[0:2])))
#
#             rt_sign = 1
#             if rays[i].raycode[j][1] == rays[i].raycode[j + 1][1]:
#                 rt_sign = - 1
#
#             l_primes[j + 1] = 1 / (1 / (l_primes[j] + l[j]) * v[j + 1] / v[j] *
#                                    cos_inc**2 / cos_out**2 +
#                                    rays[i].raycode[j][0] * 1 / cos_out *
#                                    (v[j + 1] / v[j] *
#                                     cos_inc / cos_out - rt_sign) * 1 / 2500)
#
#             lv_sum = 0
#             for k in range(j + 1):
#                 lv_sum = lv_sum + l[k] * v[k]
#
#             l_double_primes[j + 1] = 1 /( v[j + 1] / lv_sum )
#
#         theor_spread[i] = l[0] * np.sqrt((1 + l[1] / l_primes[1]) * (1 + l[1] / l_double_primes[1]) *
#                                          (1 + l[2] / l_primes[2]) * (1 + l[2] / l_double_primes[2]) *
#                                          (1 + l[3] / l_primes[3]) * (1 + l[3] / l_double_primes[3]))
#
#     fig7 = plt.figure()
#
#     plt.plot(rec_line[0:geom_spread_curv.shape[0]], np.sqrt(geom_spread_curv),label="Расчётное")
#     plt.plot(rec_line[0:theor_spread.shape[0]], theor_spread,"ro",label="Аналитическое")
#
#     plt.legend()
#     plt.grid()
#
#     plt.show()

# new_rays = np.empty(sou_line.shape[0], dtype = Ray)
# new_time = np.zeros(new_rays.shape)
#
# for i in range(new_rays.shape[0]):
#
#     new_rays[i] = Ray(sources[25], receivers[- 1 - i], current_mod, current_raycode)
#     new_rays[i].optimize(snells_law = True, dtravel = True)
#
#     new_time[i] = new_rays[i].travel_time()
#
# fig = plt.figure()
# ax = Axes3D(fig)
# ax.invert_zaxis()
#
# ax.view_init(0, - 90)
#
# for l in current_mod.layers[:-1]:
#     l.bottom.plot(ax=ax)
#
# for i in range(new_rays.shape[0]):
#
#     sources[- 1 - i].plot(ax=ax, color='r', marker='p', s=50)
#     receivers[- 1 - i].plot(ax=ax, color='k', marker='^', s=50)
#
#     new_rays[i].plot(ax=ax)
#
# # Create cubic bounding box to simulate equal aspect ratio
# max_range = np.array([horizons[-1].X.max()-horizons[-1].X.min(), horizons[-1].Y.max()-horizons[-1].Y.min(), horizons[-1].Z.max()-horizons[-1].Z.min()]).max()
# Xb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][0].flatten() + 0.5*(horizons[-1].X.max()+horizons[-1].X.min())
# Yb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][1].flatten() + 0.5*(horizons[-1].Y.max()+horizons[-1].Y.min())
# Zb = 0.5*max_range*np.mgrid[-1:2:2,-1:2:2,-1:2:2][2].flatten() + 0.5*(horizons[-1].Z.max()+horizons[-1].Z.min())
# # Comment or uncomment following both lines to test the fake bounding box:
# for xb, yb, zb in zip(Xb, Yb, Zb):
#     ax.plot([xb], [yb], [zb], 'w')
#
# ax.set_xlabel("Расстояние по оси x, м")
# # ax.set_ylabel("Расстояние по оси y, м",)
# ax.set_zlabel("Глубина, м")
#
# plt.show()
#
# plt.figure()
#
# plt.plot(rec_line, new_time,"r-")
#
# plt.title("Годограф")
#
# plt.plot(rec_line[0 : new_time.shape[0]], new_time, 'r-', label = "Годограф отражённой PP-волны")
#
# plt.legend()
# plt.grid()
#
# plt.xlabel("Координаты вдоль профиля, м")
# plt.ylabel("Время первых вступлений, с")
#
# plt.show()